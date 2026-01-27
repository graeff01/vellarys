"""
SERVIÇO DE EXPORTAÇÃO
======================

Exporta dados de leads e métricas em diferentes formatos:
- Excel (.xlsx) - Múltiplas abas com formatação profissional
- CSV (.csv) - Dados simples para importação
- PDF (.pdf) - Relatório visual para apresentações
"""

import io
import csv
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, 
    PageBreak, Image
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession

from src.domain.entities import Lead, Message, Tenant


# ============================================
# CONSTANTES E CONFIGURAÇÕES
# ============================================

# Cores do Vellarys
VELLARYS_GREEN = "4CAF50"
VELLARYS_BLUE = "2196F3"
VELLARYS_ORANGE = "FF9800"
VELLARYS_GRAY = "F5F5F5"

# Mapeamento de status para português
STATUS_MAP = {
    'new': 'Novo',
    'novo': 'Novo',
    'in_progress': 'Em atendimento',
    'em_atendimento': 'Em atendimento',
    'qualified': 'Qualificado',
    'qualificado': 'Qualificado',
    'handed_off': 'Transferido',
    'transferred': 'Transferido',
    'converted': 'Convertido',
    'convertido': 'Convertido',
    'lost': 'Perdido',
    'perdido': 'Perdido',
}

# Mapeamento de qualificação para português
QUALIFICATION_MAP = {
    'new': 'Novo', 
    'novo': 'Novo',
    'cold': 'Frio',
    'frio': 'Frio',
    'warm': 'Morno',
    'morno': 'Morno',
    'hot': 'Quente',
    'quente': 'Quente',
}


# ============================================
# FUNÇÕES DE BUSCA DE DADOS
# ============================================

async def get_leads_for_export(
    db: AsyncSession,
    tenant_id: int,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
) -> List[Dict[str, Any]]:
    """Busca leads para exportação com contagem de mensagens."""
    
    # Query base
    query = select(Lead).where(Lead.tenant_id == tenant_id)
    
    # Filtro de data
    if start_date:
        query = query.where(Lead.created_at >= start_date)
    if end_date:
        query = query.where(Lead.created_at <= end_date)
    
    # Ordena por data
    query = query.order_by(Lead.created_at.desc())
    
    result = await db.execute(query)
    leads = result.scalars().all()
    
    # Busca contagem de mensagens para cada lead
    leads_data = []
    for lead in leads:
        msg_count_result = await db.execute(
            select(func.count(Message.id)).where(Message.lead_id == lead.id)
        )
        msg_count = msg_count_result.scalar() or 0
        
        custom_data = lead.custom_data or {}
        
        leads_data.append({
            'id': lead.id,
            'created_at': lead.created_at,
            'name': lead.name or '',
            'phone': lead.phone or '',
            'email': lead.email or '',
            'city': lead.city or '',
            'status': STATUS_MAP.get(lead.status, lead.status or 'Novo'),
            'qualification': QUALIFICATION_MAP.get(lead.qualification, lead.qualification or 'Frio'),
            'source': lead.source or 'Orgânico',
            'interest': custom_data.get('interest') or custom_data.get('interest_type') or '',
            'budget': custom_data.get('budget') or custom_data.get('budget_range') or '',
            'urgency': custom_data.get('urgency') or custom_data.get('urgency_level') or '',
            'summary': lead.summary or '',
            'message_count': msg_count,
            'last_activity': lead.last_activity_at or lead.updated_at,
        })
    
    return leads_data


async def get_metrics_for_export(
    db: AsyncSession,
    tenant_id: int,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
) -> Dict[str, Any]:
    """Calcula métricas para exportação."""
    
    # Query base
    base_filter = [Lead.tenant_id == tenant_id]
    if start_date:
        base_filter.append(Lead.created_at >= start_date)
    if end_date:
        base_filter.append(Lead.created_at <= end_date)
    
    # Total de leads
    total_result = await db.execute(
        select(func.count(Lead.id)).where(and_(*base_filter))
    )
    total_leads = total_result.scalar() or 0
    
    # Por status
    status_result = await db.execute(
        select(Lead.status, func.count(Lead.id))
        .where(and_(*base_filter))
        .group_by(Lead.status)
    )
    by_status = {row[0] or 'new': row[1] for row in status_result.all()}
    
    # Por qualificação
    qual_result = await db.execute(
        select(Lead.qualification, func.count(Lead.id))
        .where(and_(*base_filter))
        .group_by(Lead.qualification)
    )
    by_qualification = {row[0] or 'cold': row[1] for row in qual_result.all()}
    
    # Por origem
    source_result = await db.execute(
        select(Lead.source, func.count(Lead.id))
        .where(and_(*base_filter))
        .group_by(Lead.source)
    )
    by_source = {row[0] or 'Orgânico': row[1] for row in source_result.all()}
    
    # Total de mensagens
    msg_result = await db.execute(
        select(func.count(Message.id))
        .join(Lead, Message.lead_id == Lead.id)
        .where(and_(*base_filter))
    )
    total_messages = msg_result.scalar() or 0
    
    # Leads convertidos
    converted = by_status.get('converted', 0) + by_status.get('convertido', 0)
    conversion_rate = (converted / total_leads * 100) if total_leads > 0 else 0
    
    # Leads quentes
    hot_leads = by_qualification.get('hot', 0) + by_qualification.get('quente', 0)
    
    return {
        'total_leads': total_leads,
        'total_messages': total_messages,
        'converted': converted,
        'conversion_rate': round(conversion_rate, 1),
        'hot_leads': hot_leads,
        'by_status': by_status,
        'by_qualification': by_qualification,
        'by_source': by_source,
        'avg_messages_per_lead': round(total_messages / total_leads, 1) if total_leads > 0 else 0,
    }


# ============================================
# EXPORTAÇÃO EXCEL
# ============================================

async def export_to_excel(
    db: AsyncSession,
    tenant: Tenant,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    include_metrics: bool = True,
    include_leads: bool = True,
) -> bytes:
    """
    Exporta dados para Excel com múltiplas abas e formatação profissional.
    
    Returns:
        bytes do arquivo Excel
    """
    
    # Busca dados
    leads = await get_leads_for_export(db, tenant.id, start_date, end_date)
    metrics = await get_metrics_for_export(db, tenant.id, start_date, end_date) if include_metrics else None
    
    # Cria workbook
    wb = Workbook()
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=VELLARYS_GREEN)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    cell_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    # ============================================
    # ABA 1: RESUMO
    # ============================================
    if include_metrics and metrics:
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        
        # Título
        ws_resumo['A1'] = f"📊 RELATÓRIO VELLARYS - {tenant.name}"
        ws_resumo['A1'].font = Font(bold=True, size=16, color=VELLARYS_GREEN)
        ws_resumo.merge_cells('A1:D1')
        
        # Período
        periodo = "Todo o período"
        if start_date and end_date:
            periodo = f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
        elif start_date:
            periodo = f"A partir de {start_date.strftime('%d/%m/%Y')}"
        
        ws_resumo['A2'] = f"Período: {periodo}"
        ws_resumo['A2'].font = Font(italic=True, color="666666")
        
        ws_resumo['A3'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
        ws_resumo['A3'].font = Font(italic=True, color="666666")
        
        # Métricas principais
        ws_resumo['A5'] = "MÉTRICAS PRINCIPAIS"
        ws_resumo['A5'].font = Font(bold=True, size=12)
        
        metricas = [
            ("Total de Leads", metrics['total_leads']),
            ("Leads Quentes 🔥", metrics['hot_leads']),
            ("Convertidos ✅", metrics['converted']),
            ("Taxa de Conversão", f"{metrics['conversion_rate']}%"),
            ("Total de Mensagens", metrics['total_messages']),
            ("Média msgs/lead", metrics['avg_messages_per_lead']),
        ]
        
        for i, (label, value) in enumerate(metricas):
            row = 6 + i
            ws_resumo[f'A{row}'] = label
            ws_resumo[f'B{row}'] = value
            ws_resumo[f'A{row}'].font = Font(bold=True)
            ws_resumo[f'B{row}'].alignment = number_alignment
        
        # Por Status
        ws_resumo['A14'] = "POR STATUS"
        ws_resumo['A14'].font = Font(bold=True, size=12)
        
        row = 15
        for status, count in metrics['by_status'].items():
            ws_resumo[f'A{row}'] = STATUS_MAP.get(status, status)
            ws_resumo[f'B{row}'] = count
            row += 1
        
        # Por Qualificação
        ws_resumo['D14'] = "POR TEMPERATURA"
        ws_resumo['D14'].font = Font(bold=True, size=12)
        
        row = 15
        for qual, count in metrics['by_qualification'].items():
            ws_resumo[f'D{row}'] = QUALIFICATION_MAP.get(qual, qual)
            ws_resumo[f'E{row}'] = count
            row += 1
        
        # Ajusta largura das colunas
        ws_resumo.column_dimensions['A'].width = 25
        ws_resumo.column_dimensions['B'].width = 15
        ws_resumo.column_dimensions['D'].width = 20
        ws_resumo.column_dimensions['E'].width = 15
    
    # ============================================
    # ABA 2: LEADS
    # ============================================
    if include_leads and leads:
        ws_leads = wb.create_sheet("Leads") if include_metrics else wb.active
        if not include_metrics:
            ws_leads.title = "Leads"
        
        # Cabeçalho
        headers = [
            "Data", "Nome", "Telefone", "Email", "Cidade", 
            "Status", "Temperatura", "Origem", "Interesse",
            "Orçamento", "Urgência", "Mensagens", "Resumo IA"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws_leads.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Dados
        for row_num, lead in enumerate(leads, 2):
            data = [
                lead['created_at'].strftime('%d/%m/%Y %H:%M') if lead['created_at'] else '',
                lead['name'],
                lead['phone'],
                lead['email'],
                lead['city'],
                lead['status'],
                lead['qualification'],
                lead['source'],
                lead['interest'],
                lead['budget'],
                lead['urgency'],
                lead['message_count'],
                lead['summary'][:100] + '...' if len(lead['summary']) > 100 else lead['summary'],
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_leads.cell(row=row_num, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                # Formatação condicional por temperatura
                if col == 7:  # Temperatura
                    if value == 'Quente':
                        cell.fill = PatternFill("solid", fgColor="FFCDD2")
                    elif value == 'Morno':
                        cell.fill = PatternFill("solid", fgColor="FFE0B2")
                    else:
                        cell.fill = PatternFill("solid", fgColor="BBDEFB")
        
        # Ajusta largura das colunas
        column_widths = [18, 25, 15, 25, 15, 15, 12, 15, 20, 15, 15, 12, 40]
        for i, width in enumerate(column_widths, 1):
            ws_leads.column_dimensions[get_column_letter(i)].width = width
        
        # Congela primeira linha
        ws_leads.freeze_panes = 'A2'
        
        # Filtros
        ws_leads.auto_filter.ref = f"A1:M{len(leads) + 1}"
    
    # Salva em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


# ============================================
# EXPORTAÇÃO CSV
# ============================================

async def export_to_csv(
    db: AsyncSession,
    tenant: Tenant,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
) -> bytes:
    """
    Exporta leads para CSV.
    
    Returns:
        bytes do arquivo CSV
    """
    
    leads = await get_leads_for_export(db, tenant.id, start_date, end_date)
    
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    
    # Cabeçalho
    headers = [
        "Data", "Nome", "Telefone", "Email", "Cidade",
        "Status", "Temperatura", "Origem", "Interesse",
        "Orçamento", "Urgência", "Mensagens", "Resumo"
    ]
    writer.writerow(headers)
    
    # Dados
    for lead in leads:
        row = [
            lead['created_at'].strftime('%d/%m/%Y %H:%M') if lead['created_at'] else '',
            lead['name'],
            lead['phone'],
            lead['email'],
            lead['city'],
            lead['status'],
            lead['qualification'],
            lead['source'],
            lead['interest'],
            lead['budget'],
            lead['urgency'],
            lead['message_count'],
            lead['summary'],
        ]
        writer.writerow(row)
    
    return buffer.getvalue().encode('utf-8-sig')  # BOM para Excel reconhecer UTF-8


# ============================================
# EXPORTAÇÃO PDF
# ============================================

async def export_to_pdf(
    db: AsyncSession,
    tenant: Tenant,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    include_leads: bool = True,
    max_leads: int = 50,
) -> bytes:
    """
    Exporta relatório para PDF com visual profissional.
    
    Returns:
        bytes do arquivo PDF
    """
    
    # Busca dados
    leads = await get_leads_for_export(db, tenant.id, start_date, end_date)
    metrics = await get_metrics_for_export(db, tenant.id, start_date, end_date)
    
    # Cria buffer
    buffer = io.BytesIO()
    
    # Cria documento
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4CAF50'),
        alignment=TA_CENTER,
        spaceAfter=20,
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#666666'),
        alignment=TA_CENTER,
        spaceAfter=30,
    )
    
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2196F3'),
        spaceBefore=20,
        spaceAfter=10,
    )
    
    # Conteúdo
    story = []
    
    # Título
    story.append(Paragraph(f"📊 Relatório Vellarys", title_style))
    story.append(Paragraph(f"{tenant.name}", subtitle_style))
    
    # Período
    periodo = "Todo o período"
    if start_date and end_date:
        periodo = f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
    
    story.append(Paragraph(f"Período: {periodo}", styles['Normal']))
    story.append(Paragraph(
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", 
        styles['Normal']
    ))
    story.append(Spacer(1, 30))
    
    # Métricas principais
    story.append(Paragraph("📈 Métricas Principais", section_style))
    
    metrics_data = [
        ['Métrica', 'Valor'],
        ['Total de Leads', str(metrics['total_leads'])],
        ['Leads Quentes 🔥', str(metrics['hot_leads'])],
        ['Convertidos ✅', str(metrics['converted'])],
        ['Taxa de Conversão', f"{metrics['conversion_rate']}%"],
        ['Total de Mensagens', str(metrics['total_messages'])],
        ['Média msgs/lead', str(metrics['avg_messages_per_lead'])],
    ]
    
    metrics_table = Table(metrics_data, colWidths=[10*cm, 5*cm])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5F5')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E0E0E0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    
    story.append(metrics_table)
    story.append(Spacer(1, 30))
    
    # Por Status
    story.append(Paragraph("📊 Leads por Status", section_style))
    
    status_data = [['Status', 'Quantidade']]
    for status, count in metrics['by_status'].items():
        status_data.append([STATUS_MAP.get(status, status), str(count)])
    
    status_table = Table(status_data, colWidths=[10*cm, 5*cm])
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E0E0E0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    
    story.append(status_table)
    story.append(Spacer(1, 20))
    
    # Por Temperatura
    story.append(Paragraph("🌡️ Leads por Temperatura", section_style))
    
    qual_data = [['Temperatura', 'Quantidade']]
    for qual, count in metrics['by_qualification'].items():
        qual_data.append([QUALIFICATION_MAP.get(qual, qual), str(count)])
    
    qual_table = Table(qual_data, colWidths=[10*cm, 5*cm])
    qual_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF9800')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E0E0E0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    
    story.append(qual_table)
    
    # Lista de leads (se incluir)
    if include_leads and leads:
        story.append(PageBreak())
        story.append(Paragraph("📋 Lista de Leads", section_style))
        story.append(Paragraph(
            f"Mostrando {min(len(leads), max_leads)} de {len(leads)} leads",
            styles['Normal']
        ))
        story.append(Spacer(1, 10))
        
        # Tabela de leads (simplificada para caber no PDF)
        leads_header = ['Nome', 'Telefone', 'Status', 'Temp.', 'Msgs']
        leads_data = [leads_header]
        
        for lead in leads[:max_leads]:
            leads_data.append([
                lead['name'][:20] if lead['name'] else '-',
                lead['phone'] or '-',
                lead['status'],
                lead['qualification'],
                str(lead['message_count']),
            ])
        
        leads_table = Table(leads_data, colWidths=[5*cm, 4*cm, 3*cm, 2.5*cm, 1.5*cm])
        leads_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        
        story.append(leads_table)
    
    # Rodapé
    story.append(Spacer(1, 40))
    story.append(Paragraph(
        "Relatório gerado automaticamente pelo Vellarys - Sistema de Atendimento com IA",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray, alignment=TA_CENTER)
    ))
    
    # Build
    doc.build(story)
    
    buffer.seek(0)
    return buffer.getvalue()